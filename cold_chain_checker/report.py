import csv
import os
from collections import defaultdict

from cold_chain_checker.models import ValidationIssue


CATEGORY_LABELS = {
    "departure_time": "发车时间",
    "arrival_location": "到达地点",
    "temperature_continuity": "温度断点",
    "temperature_range": "温度超限",
    "receipt_signature": "签收签字",
    "receipt_box_check": "箱码核对",
    "receipt_box_photo": "箱码照片",
    "load_failed": "加载失败",
}

CATEGORY_ORDER = [
    "load_failed",
    "temperature_continuity",
    "temperature_range",
    "departure_time",
    "arrival_location",
    "receipt_signature",
    "receipt_box_check",
    "receipt_box_photo",
]

RECTIFICATION_BUCKETS = {
    "待补数据": {
        "desc": "数据文件缺失或格式损坏，需重新补数据",
        "categories": {"load_failed"},
    },
    "待司机补签": {
        "desc": "签收资料不全，需司机补签字、补拍照片",
        "categories": {"receipt_signature", "receipt_box_check", "receipt_box_photo"},
    },
    "待质控复核": {
        "desc": "冷链过程异常，需质控复核并填写异常说明",
        "categories": {"temperature_continuity", "temperature_range", "departure_time", "arrival_location"},
    },
}


def _summarize_issues(issues: list) -> str:
    if not issues:
        return ""
    parts = []
    for issue in issues[:3]:
        label = CATEGORY_LABELS.get(issue.category, issue.category)
        parts.append(f"[{label}] {issue.message}")
    if len(issues) > 3:
        parts.append(f"...等共 {len(issues)} 项")
    return "；".join(parts)


def _build_vehicle_summary_rows(results: list) -> list:
    vehicle_data = defaultdict(lambda: {
        "total": 0,
        "passed": 0,
        "failed": 0,
        "load_failed": 0,
        "routes": set(),
        "carriers": set(),
        "issues": [],
    })

    for item in results:
        waybill = item.get("waybill")
        is_load_failed = item.get("load_failed", False)

        if not waybill:
            if is_load_failed:
                key = "（未知车辆）"
                entry = vehicle_data[key]
                entry["total"] += 1
                entry["load_failed"] += 1
                entry["failed"] += 1
                for issue in item.get("issues", []):
                    entry["issues"].append(issue)
            continue

        key = waybill.vehicle_plate
        entry = vehicle_data[key]
        entry["total"] += 1
        entry["routes"].add(waybill.route)
        entry["carriers"].add(waybill.carrier)

        if is_load_failed:
            entry["load_failed"] += 1
            entry["failed"] += 1
            for issue in item.get("issues", []):
                entry["issues"].append(issue)
        else:
            issues = item.get("issues", [])
            if issues:
                entry["failed"] += 1
                for issue in issues:
                    entry["issues"].append(issue)
            else:
                entry["passed"] += 1

    rows = []
    for vehicle in sorted(vehicle_data.keys()):
        data = vehicle_data[vehicle]
        route_str = "、".join(sorted(data["routes"]))
        carrier_str = "、".join(sorted(data["carriers"]))
        error_count = sum(1 for i in data["issues"] if i.severity == "error")
        warning_count = sum(1 for i in data["issues"] if i.severity == "warning")

        by_category = defaultdict(list)
        for issue in data["issues"]:
            by_category[issue.category].append(issue)

        summary_parts = []
        for cat in CATEGORY_ORDER:
            if cat not in by_category:
                continue
            cat_items = by_category[cat]
            label = CATEGORY_LABELS.get(cat, cat)
            summary_parts.append(f"{label} {len(cat_items)} 条")
        main_issues = "，".join(summary_parts) if summary_parts else "无"

        rows.append({
            "承运车辆": vehicle,
            "承运商": carrier_str,
            "当天运单数": data["total"],
            "通过": data["passed"],
            "异常": data["failed"],
            "其中加载失败": data["load_failed"],
            "错误数": error_count,
            "警告数": warning_count,
            "运输线路": route_str,
            "主要问题": main_issues,
        })

    return rows


def _build_anomaly_type_rows(results: list) -> list:
    category_data = defaultdict(lambda: {})

    for item in results:
        waybill = item.get("waybill")
        issues = item.get("issues", [])
        if not issues:
            continue

        if waybill:
            wb_id = waybill.waybill_id
            vehicle = waybill.vehicle_plate
            route = waybill.route
            products = "、".join(sorted(set(b.product for b in waybill.vaccine_boxes)))
        else:
            wb_id = item.get("folder_name", "未知")
            vehicle = "未知"
            route = "未知"
            products = "未知"

        for issue in issues:
            cat = issue.category
            cat_label = CATEGORY_LABELS.get(cat, cat)
            if wb_id not in category_data[cat]:
                category_data[cat][wb_id] = {
                    "vehicle": vehicle,
                    "route": route,
                    "products": products,
                    "messages": [],
                }
            category_data[cat][wb_id]["messages"].append(issue.message)

    rows = []
    for cat in CATEGORY_ORDER:
        if cat not in category_data:
            continue
        cat_label = CATEGORY_LABELS.get(cat, cat)
        for wb_id in sorted(category_data[cat].keys()):
            info = category_data[cat][wb_id]
            rows.append({
                "异常类型": cat_label,
                "运单号": wb_id,
                "承运车辆": info["vehicle"],
                "线路": info["route"],
                "疫苗品种": info["products"],
                "异常条数": len(info["messages"]),
                "异常详情": "；".join(info["messages"]),
            })

    return rows


def _build_unmatched_rows(results: list) -> list:
    rows = []
    for item in results:
        if not item.get("filter_unmatched", False):
            continue
        rows.append({
            "文件夹名": item.get("folder_name", "未知"),
            "失败原因": item.get("error_message", ""),
            "说明": "waybill.json 无法读取，无法判断是否属于当前筛选条件",
        })
    return rows


def _build_product_liability_rows(results: list) -> list:
    product_data = defaultdict(lambda: {
        "vehicles": set(),
        "routes": set(),
        "carriers": set(),
        "waybill_details": {},
    })

    for item in results:
        waybill = item.get("waybill")
        if not waybill:
            continue
        issues = item.get("issues", [])

        unique_products_in_waybill = set()
        for box in waybill.vaccine_boxes:
            unique_products_in_waybill.add(box.product)

        for product in unique_products_in_waybill:
            entry = product_data[product]
            entry["vehicles"].add(waybill.vehicle_plate)
            entry["routes"].add(waybill.route)
            entry["carriers"].add(waybill.carrier)

            if waybill.waybill_id not in entry["waybill_details"]:
                by_category = defaultdict(list)
                for issue in issues:
                    label = CATEGORY_LABELS.get(issue.category, issue.category)
                    by_category[label].append(issue.message)

                anomaly_parts = []
                for cat in [CATEGORY_LABELS[c] for c in CATEGORY_ORDER if c in CATEGORY_LABELS]:
                    if cat in by_category:
                        anomaly_parts.append(f"{cat} {len(by_category[cat])}")

                entry["waybill_details"][waybill.waybill_id] = {
                    "vehicle": waybill.vehicle_plate,
                    "route": waybill.route,
                    "carrier": waybill.carrier,
                    "status": "异常" if issues else "通过",
                    "anomaly_types": "，".join(anomaly_parts) if anomaly_parts else "无",
                    "anomaly_detail": _summarize_issues(issues),
                }

    rows = []
    for product in sorted(product_data.keys()):
        data = product_data[product]
        total = len(data["waybill_details"])
        passed = sum(1 for d in data["waybill_details"].values() if d["status"] == "通过")
        failed = total - passed
        vehicle_str = "、".join(sorted(data["vehicles"]))
        carrier_str = "、".join(sorted(data["carriers"]))
        route_str = "、".join(sorted(data["routes"]))

        for wb_id in sorted(data["waybill_details"].keys()):
            info = data["waybill_details"][wb_id]
            rows.append({
                "疫苗品种": product,
                "品种总单数": total,
                "品种通过": passed,
                "品种异常": failed,
                "品种涉及车辆": vehicle_str,
                "品种涉及承运商": carrier_str,
                "品种涉及线路": route_str,
                "运单号": wb_id,
                "承运车辆": info["vehicle"],
                "线路": info["route"],
                "承运商": info["carrier"],
                "通过状态": info["status"],
                "异常类型": info["anomaly_types"],
                "异常详情": info["anomaly_detail"],
            })

    return rows


def _build_rectification_rows(results: list) -> list:
    rows = []
    for bucket_name in ["待补数据", "待司机补签", "待质控复核"]:
        bucket_def = RECTIFICATION_BUCKETS[bucket_name]
        bucket_cats = bucket_def["categories"]

        bucket_waybills = {}
        for item in results:
            waybill = item.get("waybill")
            issues = item.get("issues", [])
            if not issues:
                continue

            if waybill:
                wb_id = waybill.waybill_id
                vehicle = waybill.vehicle_plate
                route = waybill.route
                carrier = waybill.carrier
            else:
                wb_id = item.get("folder_name", "未知")
                vehicle = "未知"
                route = "未知"
                carrier = "未知"

            relevant = [i for i in issues if i.category in bucket_cats]
            if not relevant:
                continue

            if wb_id not in bucket_waybills:
                by_category = defaultdict(list)
                for issue in relevant:
                    label = CATEGORY_LABELS.get(issue.category, issue.category)
                    by_category[label].append(issue.message)

                cat_parts = []
                for cat in [CATEGORY_LABELS[c] for c in CATEGORY_ORDER if c in CATEGORY_LABELS]:
                    if cat in by_category:
                        cat_parts.append(f"{cat} {len(by_category[cat])}")

                all_messages = []
                for issue in relevant:
                    all_messages.append(issue.message)

                bucket_waybills[wb_id] = {
                    "vehicle": vehicle,
                    "route": route,
                    "carrier": carrier,
                    "relevant_count": len(relevant),
                    "anomaly_types": "，".join(cat_parts) if cat_parts else "无",
                    "anomaly_detail": "；".join(all_messages),
                }

        for wb_id in sorted(bucket_waybills.keys()):
            info = bucket_waybills[wb_id]
            rows.append({
                "整改分类": bucket_name,
                "分类说明": bucket_def["desc"],
                "运单号/文件夹": wb_id,
                "承运车辆": info["vehicle"],
                "线路": info["route"],
                "承运商": info["carrier"],
                "待办条数": info["relevant_count"],
                "异常类型": info["anomaly_types"],
                "待办详情": info["anomaly_detail"],
            })

    return rows


def _build_liability_rows(results: list) -> list:
    carrier_stats = defaultdict(lambda: {
        "total": 0,
        "passed": 0,
        "failed": 0,
        "load_failed": 0,
        "vehicles": set(),
        "routes": set(),
        "anomaly_counts": defaultdict(int),
    })

    for item in results:
        waybill = item.get("waybill")
        issues = item.get("issues", [])
        is_load_failed = item.get("load_failed", False)

        if waybill:
            carrier = waybill.carrier
            vehicle = waybill.vehicle_plate
            route = waybill.route
        else:
            carrier = "（未知承运商）"
            vehicle = "（未知车辆）"
            route = "（未知线路）"

        entry = carrier_stats[carrier]
        entry["total"] += 1
        entry["vehicles"].add(vehicle)
        entry["routes"].add(route)

        if is_load_failed:
            entry["failed"] += 1
            entry["load_failed"] += 1
        else:
            if issues:
                entry["failed"] += 1
            else:
                entry["passed"] += 1

        for issue in issues:
            label = CATEGORY_LABELS.get(issue.category, issue.category)
            entry["anomaly_counts"][label] += 1

    sorted_carriers = sorted(
        carrier_stats.items(),
        key=lambda kv: (kv[1]["failed"], kv[1]["total"]),
        reverse=True,
    )

    rows = []
    for rank, (carrier, data) in enumerate(sorted_carriers, 1):
        rate = (data["failed"] / data["total"] * 100) if data["total"] > 0 else 0

        anomaly_parts = []
        for cat in [CATEGORY_LABELS[c] for c in CATEGORY_ORDER if c in CATEGORY_LABELS]:
            if data["anomaly_counts"][cat] > 0:
                anomaly_parts.append(f"{cat} {data['anomaly_counts'][cat]}")

        rows.append({
            "排名": rank,
            "承运商": carrier,
            "总单数": data["total"],
            "通过": data["passed"],
            "异常": data["failed"],
            "其中加载失败": data["load_failed"],
            "异常率": f"{rate:.0f}%",
            "涉及车辆": "、".join(sorted(data["vehicles"])),
            "涉及线路": "、".join(sorted(data["routes"])),
            "异常明细": "；".join(anomaly_parts) if anomaly_parts else "无",
        })

    return rows


def generate_daily_report_csv(
    results: list,
    output_path: str,
) -> str:
    detail_headers = [
        "序号", "运单号/文件夹", "承运车辆", "线路", "承运商", "状态", "错误数", "警告数", "主要异常",
    ]

    rows = []
    for idx, item in enumerate(results, 1):
        waybill = item.get("waybill")
        is_unmatched = item.get("filter_unmatched", False)
        if item.get("load_failed"):
            status = "加载失败（无法匹配筛选）" if is_unmatched else "加载失败"
            row = {
                "序号": idx,
                "运单号/文件夹": item["folder_name"],
                "承运车辆": waybill.vehicle_plate if waybill else "-",
                "线路": waybill.route if waybill else "-",
                "承运商": waybill.carrier if waybill else "-",
                "状态": status,
                "错误数": 1,
                "警告数": 0,
                "主要异常": item["error_message"],
            }
        else:
            issues = item["issues"]
            error_count = sum(1 for i in issues if i.severity == "error")
            warning_count = sum(1 for i in issues if i.severity == "warning")
            status = "通过" if not issues else "异常"
            row = {
                "序号": idx,
                "运单号/文件夹": waybill.waybill_id,
                "承运车辆": waybill.vehicle_plate,
                "线路": waybill.route,
                "承运商": waybill.carrier,
                "状态": status,
                "错误数": error_count,
                "警告数": warning_count,
                "主要异常": _summarize_issues(issues),
            }
        rows.append(row)

    vehicle_headers = [
        "承运车辆", "承运商", "当天运单数", "通过", "异常", "其中加载失败", "错误数", "警告数", "运输线路", "主要问题",
    ]
    vehicle_rows = _build_vehicle_summary_rows(results)

    anomaly_headers = [
        "异常类型", "运单号", "承运车辆", "线路", "疫苗品种", "异常条数", "异常详情",
    ]
    anomaly_rows = _build_anomaly_type_rows(results)

    product_headers = [
        "疫苗品种", "品种总单数", "品种通过", "品种异常", "品种涉及车辆", "品种涉及承运商", "品种涉及线路",
        "运单号", "承运车辆", "线路", "承运商", "通过状态", "异常类型", "异常详情",
    ]
    product_rows = _build_product_liability_rows(results)

    rectification_headers = [
        "整改分类", "分类说明", "运单号/文件夹", "承运车辆", "线路", "承运商", "待办条数", "异常类型", "待办详情",
    ]
    rectification_rows = _build_rectification_rows(results)

    liability_headers = [
        "排名", "承运商", "总单数", "通过", "异常", "其中加载失败", "异常率", "涉及车辆", "涉及线路", "异常明细",
    ]
    liability_rows = _build_liability_rows(results)

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=detail_headers)
        writer.writeheader()
        writer.writerows(rows)

        f.write("\n\n")
        f.write("=== 车辆分组小结 ===\n")
        writer2 = csv.DictWriter(f, fieldnames=vehicle_headers)
        writer2.writeheader()
        writer2.writerows(vehicle_rows)

        if anomaly_rows:
            f.write("\n\n")
            f.write("=== 异常类型汇总 ===\n")
            writer3 = csv.DictWriter(f, fieldnames=anomaly_headers)
            writer3.writeheader()
            writer3.writerows(anomaly_rows)

        if product_rows:
            f.write("\n\n")
            f.write("=== 品种追责台账 ===\n")
            writer5 = csv.DictWriter(f, fieldnames=product_headers)
            writer5.writeheader()
            writer5.writerows(product_rows)

        if rectification_rows:
            f.write("\n\n")
            f.write("=== 整改跟踪清单 ===\n")
            writer6 = csv.DictWriter(f, fieldnames=rectification_headers)
            writer6.writeheader()
            writer6.writerows(rectification_rows)

        if liability_rows:
            f.write("\n\n")
            f.write("=== 追责责任口径小结 ===\n")
            writer7 = csv.DictWriter(f, fieldnames=liability_headers)
            writer7.writeheader()
            writer7.writerows(liability_rows)

        unmatched_rows = _build_unmatched_rows(results)
        if unmatched_rows:
            unmatched_headers = ["文件夹名", "失败原因", "说明"]
            f.write("\n\n")
            f.write("=== 补数据区域（无法匹配筛选条件） ===\n")
            writer4 = csv.DictWriter(f, fieldnames=unmatched_headers)
            writer4.writeheader()
            writer4.writerows(unmatched_rows)

    return output_path


def _write_xlsx_sheet(ws, headers, rows, header_font, header_fill, center_align, left_align, thin_border, center_cols=None):
    center_cols = center_cols or set()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "") if isinstance(row_data, dict) else row_data[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in center_cols:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.row_dimensions[1].height = 28


def _set_column_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)
        ws.column_dimensions[col_letter].width = width


def generate_daily_report_excel(
    results: list,
    output_path: str,
) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("未安装 openpyxl，请先安装：pip install openpyxl")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    load_fail_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    unmatched_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    detail_headers = [
        "序号", "运单号/文件夹", "承运车辆", "线路", "承运商", "状态", "错误数", "警告数", "主要异常",
    ]

    ws = wb.active
    ws.title = "运单明细"
    _write_xlsx_sheet(ws, detail_headers, [], header_font, header_fill,
                      center_align, left_align, thin_border, {1, 3, 6, 7, 8})

    for idx, item in enumerate(results, 2):
        waybill = item.get("waybill")
        is_unmatched = item.get("filter_unmatched", False)
        if item.get("load_failed"):
            status = "加载失败（无法匹配筛选）" if is_unmatched else "加载失败"
            row_data = [
                idx - 1,
                item["folder_name"],
                waybill.vehicle_plate if waybill else "-",
                waybill.route if waybill else "-",
                waybill.carrier if waybill else "-",
                status,
                1,
                0,
                item["error_message"],
            ]
            status_fill = unmatched_fill if is_unmatched else load_fail_fill
        else:
            issues = item["issues"]
            error_count = sum(1 for i in issues if i.severity == "error")
            warning_count = sum(1 for i in issues if i.severity == "warning")
            status = "通过" if not issues else "异常"
            row_data = [
                idx - 1,
                waybill.waybill_id,
                waybill.vehicle_plate,
                waybill.route,
                waybill.carrier,
                status,
                error_count,
                warning_count,
                _summarize_issues(issues),
            ]
            status_fill = pass_fill if not issues else fail_fill

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 6:
                cell.fill = status_fill
            if col_idx in (1, 3, 6, 7, 8):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    _set_column_widths(ws, [6, 22, 14, 28, 12, 22, 8, 8, 50])

    vehicle_headers = [
        "承运车辆", "承运商", "当天运单数", "通过", "异常", "其中加载失败", "错误数", "警告数", "运输线路", "主要问题",
    ]
    vehicle_rows = _build_vehicle_summary_rows(results)
    ws2 = wb.create_sheet(title="车辆分组小结")
    _write_xlsx_sheet(ws2, vehicle_headers, [v for v in vehicle_rows], header_font, header_fill,
                      center_align, left_align, thin_border, {3, 4, 5, 6, 7, 8})
    _set_column_widths(ws2, [14, 14, 12, 8, 8, 14, 8, 8, 40, 50])

    anomaly_headers = [
        "异常类型", "运单号", "承运车辆", "线路", "疫苗品种", "异常条数", "异常详情",
    ]
    anomaly_rows = _build_anomaly_type_rows(results)
    if anomaly_rows:
        ws3 = wb.create_sheet(title="异常类型汇总")
        _write_xlsx_sheet(ws3, anomaly_headers, anomaly_rows, header_font, header_fill,
                          center_align, left_align, thin_border, {6})
        _set_column_widths(ws3, [12, 22, 14, 28, 30, 10, 60])

    product_headers = [
        "疫苗品种", "品种总单数", "品种通过", "品种异常", "品种涉及车辆", "品种涉及承运商", "品种涉及线路",
        "运单号", "承运车辆", "线路", "承运商", "通过状态", "异常类型", "异常详情",
    ]
    product_rows = _build_product_liability_rows(results)
    if product_rows:
        ws5 = wb.create_sheet(title="品种追责台账")
        _write_xlsx_sheet(ws5, product_headers, product_rows, header_font, header_fill,
                          center_align, left_align, thin_border, {2, 3, 4, 12})
        _set_column_widths(ws5, [20, 10, 8, 8, 26, 14, 30, 22, 14, 28, 14, 10, 30, 60])

    rectification_headers = [
        "整改分类", "分类说明", "运单号/文件夹", "承运车辆", "线路", "承运商", "待办条数", "异常类型", "待办详情",
    ]
    rectification_rows = _build_rectification_rows(results)
    if rectification_rows:
        ws6 = wb.create_sheet(title="整改跟踪清单")
        _write_xlsx_sheet(ws6, rectification_headers, rectification_rows, header_font, header_fill,
                          center_align, left_align, thin_border, {7})
        _set_column_widths(ws6, [12, 40, 22, 14, 28, 14, 10, 30, 60])

    liability_headers = [
        "排名", "承运商", "总单数", "通过", "异常", "其中加载失败", "异常率", "涉及车辆", "涉及线路", "异常明细",
    ]
    liability_rows = _build_liability_rows(results)
    if liability_rows:
        ws7 = wb.create_sheet(title="追责责任小结")
        _write_xlsx_sheet(ws7, liability_headers, liability_rows, header_font, header_fill,
                          center_align, left_align, thin_border, {1, 3, 4, 5, 6, 7})
        _set_column_widths(ws7, [8, 16, 10, 8, 8, 14, 10, 26, 30, 60])

    unmatched_rows = _build_unmatched_rows(results)
    if unmatched_rows:
        unmatched_headers = ["文件夹名", "失败原因", "说明"]
        ws4 = wb.create_sheet(title="补数据区域")
        _write_xlsx_sheet(ws4, unmatched_headers, unmatched_rows, header_font, header_fill,
                          center_align, left_align, thin_border)
        _set_column_widths(ws4, [24, 50, 50])

    wb.save(output_path)
    return output_path


def export_daily_report(
    results: list,
    output_path: str,
) -> str:
    ext = os.path.splitext(output_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return generate_daily_report_excel(results, output_path)
    else:
        return generate_daily_report_csv(results, output_path)
